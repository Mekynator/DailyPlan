import os
import logging
from flask import Flask, url_for, render_template
from dotenv import load_dotenv
from office365.runtime.auth.user_credential import UserCredential
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

app = Flask(__name__)

# Load environment variables
load_dotenv()

SHAREPOINT_SITE_URL = os.getenv('SHAREPOINT_SITE_URL')
SHAREPOINT_FILE_URL = os.getenv('SHAREPOINT_FILE_URL')
SHAREPOINT_USERNAME = os.getenv('SHAREPOINT_USERNAME')
SHAREPOINT_PASSWORD = os.getenv('SHAREPOINT_PASSWORD')

# Logger setup
logger = logging.getLogger('FlaskAppLogger')
logger.setLevel(logging.INFO)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Log to a persistent directory
LOG_FILE_PATH = os.path.join(app.root_path, 'logs', 'error.log')
os.makedirs(os.path.dirname(LOG_FILE_PATH), exist_ok=True)
file_handler = logging.FileHandler(LOG_FILE_PATH)
file_handler.setLevel(logging.ERROR)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

logger.addHandler(console_handler)
logger.addHandler(file_handler)

logger.info("Logger initialized.")

# Image storage
IMAGE_FOLDER = os.path.join(app.root_path, 'static', 'images')
os.makedirs(IMAGE_FOLDER, exist_ok=True)

def download_sharepoint_file():
    """Download the Excel file from SharePoint"""
    try:
        # Verify that all environment variables are set
        if not all([SHAREPOINT_SITE_URL, SHAREPOINT_FILE_URL, SHAREPOINT_USERNAME, SHAREPOINT_PASSWORD]):
            logger.error("One or more SharePoint environment variables are not set.")
            raise ValueError("SharePoint configuration incomplete.")
        
        # Set up SharePoint authentication
        credentials = UserCredential(SHAREPOINT_USERNAME, SHAREPOINT_PASSWORD)
        ctx = ClientContext(SHAREPOINT_SITE_URL).with_credentials(credentials)
        
        # Fetch the file from SharePoint
        file = ctx.web.get_file_by_server_relative_url(SHAREPOINT_FILE_URL)
        ctx.load(file)
        ctx.execute_query()
        
        # Use tempfile to create a temporary file for the downloaded Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as temp_file:
            file.download(temp_file)
            ctx.execute_query()
            temp_file_path = temp_file.name
        
        logger.info(f"File downloaded successfully to {temp_file_path}")
        return temp_file_path
    except Exception as e:
        logger.error(f"Failed to download SharePoint file: {e}")
        return None

def generate_image(sheet_name, cell_range, image_path):
    """Generate an image from an Excel sheet using openpyxl and matplotlib"""
    try:
        # Download the file from SharePoint
        excel_file_path = download_sharepoint_file()
        if not excel_file_path:
            raise Exception("Unable to download Excel file from SharePoint.")
        
        # Load workbook
        workbook = load_workbook_simple(excel_file_path)
        if not workbook:
            raise Exception("Unable to load workbook for image export.")
        
        # Check if the sheet exists
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        
        # Load the sheet
        wb_sheet = workbook[sheet_name]

        # Extract the data from the specified cell range (e.g., 'A1:H33')
        cells = wb_sheet[cell_range]
        data = []
        for row in cells:
            data_row = [cell.value for cell in row]
            data.append(data_row)

        # Convert the data to a numpy array for plotting
        data_array = np.array(data)

        # Plot the data as an image using matplotlib
        fig, ax = plt.subplots(figsize=(10, 6))  # Set the size of the plot
        ax.axis('off')  # Turn off the axis
        ax.table(cellText=data, cellLoc='center', loc='center', colLabels=None, cellColours=None, bbox=[0, 0, 1, 1])

        # Save the plot as an image
        plt.savefig(image_path, bbox_inches='tight', pad_inches=0.1)
        plt.close(fig)  # Close the figure to free memory

        logger.info(f"Generated image for '{sheet_name}' -> {image_path}")

    except Exception as e:
        logger.error(f"Error generating image for '{sheet_name}': {e}")
        raise e


@app.route('/')
def home():
    """Main route to render the web page"""
    active_pages = ['Morning', 'Evening', 'Night', 'Friday']
    page_ranges = {
        'Morning': ('Morning', 'A1:H33'),
        'Evening': ('Evening', 'A1:H33'),
        'Night': ('Night', 'A1:H33'),
        'Friday': ('Friday', 'A1:H33')
    }

    pages = []
    for page in active_pages:
        if page not in page_ranges:
            logger.error(f"No cell range defined for page '{page}'. Skipping.")
            continue

        sheet_name, cell_range = page_ranges[page]
        image_filename = f"{page}.png"
        image_path = os.path.join(IMAGE_FOLDER, image_filename)

        # Check if the image already exists to avoid regenerating it
        if not os.path.exists(image_path):
            try:
                generate_image(sheet_name, cell_range, image_path)
            except Exception as e:
                logger.error(f"Failed to generate image for page '{page}': {e}")
                continue

        image_url = url_for('static', filename=f'images/{image_filename}')
        pages.append({'url': image_url, 'header': f"{sheet_name} Shift"})

    if not pages:
        return "<h1>No images available for the active pages.</h1>"

    return render_template('index.html', pages=pages)

if __name__ == '__main__':
    app.run(debug=True)







































































































































































































































































































































































































































































































































































































































































































































